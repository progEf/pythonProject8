import os
import openpyxl
import pandas as pd

FIlL = input("Укажите папку: ")
# FIlL = 'C:\\Users\\mrefi\\Desktop\\проекты\\pythonProject8'
FIlLE = f'{FIlL}\\files'

file_names = input("Название файла excel: ")
# https://www.avito.ru/user/7fcc6357d8949bbceb2396ecc2eb2b2b/profile?src=search_seller_info
count = []
user = []
profile = []
alls = []
dell = []
for namb, file_name in enumerate(os.listdir(FIlLE)):
    if not file_name.startswith('~$'):  # Skip temporary files
        workbook = openpyxl.load_workbook(f'{FIlLE}\\{file_name}')
        sheet = workbook.active
        n = 0
        clear_mass = []
        profile_clear = []
        values_clear = []

        for row in sheet.iter_rows(values_only=True):
            if n == 0:
                n += 1
            else:
                #values = {'count': row[0], 'user': row[1], 'profile': row[2]}
                try:
                    values_clear = {'profile': row[2]}
                except:
                    values_clear = {'profile': row[0]}

                if values_clear not in clear_mass:
                    clear_mass.append(values_clear)
                    try:
                        profile_clear.append(row[2])
                    except:
                        profile_clear.append(row[0])

        df = pd.DataFrame(
            {
                'Ссылка на профиль продавца': profile_clear}
        )
        dell.append(f'{FIlLE}\\{file_name}')
        df.to_excel(f'{FIlL}\\files\\1{file_name}', index=False)

for i  in dell:
    os.remove(i)
for file_name in os.listdir(FIlLE):
    if not file_name.startswith('~$'):  # Skip temporary files
        workbook = openpyxl.load_workbook(f'{FIlLE}\\{file_name}')
        sheet = workbook.active
        n = 0

        for row in sheet.iter_rows(values_only=True):
            if n == 0:
                n += 1
            else:
                #values = {'count': row[0], 'user': row[1], 'profile': row[2]}
                values = {'profile': row[0]}

                if values not in alls:
                    alls.append(values)

                    profile.append(row[0])

df = pd.DataFrame(
    {
     'Ссылка на профиль продавца': profile}
)
df.to_excel(f'{FIlL}/{file_names}.xlsx', index=False)